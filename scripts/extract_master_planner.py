"""
Extract every leg from the MLC Transport Master Planner spreadsheet into a
JSON file the portal can ingest.

Reads:
  - All Mon..Sun_WK<n>_<yy> sheets         (the daily transport sheets)
  - The bottom-of-sheet "Vehicle Earnings UK Deliveries" matrix (revenue)
  - Per-week Invoicing_WK<n>_<yy> sheets   (billable + sent state)

Writes a single JSON file: an array of leg objects matching the portal's
PlannedRun shape (camelCase). The TS importer (scripts/import-master-planner.ts)
takes that JSON and upserts it to Supabase.

Pink fill (FFFDA1CF) = outbound (runType "regular")
Blue fill (FF2FC9FF) = backload  (runType "backload")

Usage:
  python scripts/extract_master_planner.py --week 39 --year 25 --out tmp/wk39.json
  python scripts/extract_master_planner.py --all --out tmp/all-weeks.json
"""

import argparse
import json
import re
import sys
import io
from pathlib import Path
from datetime import datetime, date
from typing import Optional, Iterable

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_PATH = r"C:/Users/harri/OneDrive/MLC/MLC Transport Master Planner.xlsm"

PINK_FILL = "FFFDA1CF"   # outbound
BLUE_FILL = "FF2FC9FF"   # backload

# 1-based column positions per the inspection dump.
COL_COLLECTION    = 1   # A
COL_DELIVERY      = 4   # D
COL_DAY_INDEX     = 7   # G
COL_DAY_OF        = 8   # H ("OF")
COL_DAY_COUNT     = 9   # I
COL_FACTORY       = 10  # J
COL_BOOKING_TIME  = 13  # M
COL_VEHICLE       = 14  # N
COL_SUBBY_DRIVER  = 15  # O
COL_SUBBY_COST    = 16  # P
COL_TRAILER_NUM   = 18  # R
COL_TRAILER_DROP  = 19  # S
COL_REFERENCE     = 20  # T
COL_CUSTOMER      = 22  # V


def fill_rgb(cell) -> Optional[str]:
    """Return the cell's foreground fill as an 8-char ARGB string, or None."""
    f = cell.fill
    if not f or f.fill_type is None:
        return None
    fg = f.fgColor
    if not fg:
        return None
    if fg.type == "rgb" and fg.rgb:
        return fg.rgb.upper()
    return None


def detect_run_type(row_cells) -> Optional[str]:
    """
    Walk the row's cells; return 'regular' if pink dominates, 'backload' if
    blue dominates, None otherwise. Most legs have the same fill across all
    their cells.
    """
    pink = 0
    blue = 0
    for cell in row_cells:
        rgb = fill_rgb(cell)
        if rgb == PINK_FILL:
            pink += 1
        elif rgb == BLUE_FILL:
            blue += 1
    if pink == 0 and blue == 0:
        return None
    return "backload" if blue > pink else "regular"


def safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float)):
        # Reject obvious "0" placeholder values that the spreadsheet uses
        # for empty reference / factory cells.
        if v == 0:
            return ""
        if v == int(v):
            return str(int(v))
        return str(v)
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def normalise_upper(v: str) -> str:
    """Trim, collapse whitespace, uppercase. Used for customer + vehicle."""
    return re.sub(r"\s+", " ", (v or "").strip().upper())


def safe_number(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def parse_int(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        n = int(v)
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def header_date(ws) -> Optional[str]:
    """
    Daily sheets store the date in C1. Returns yyyy-mm-dd or None.
    """
    raw = ws.cell(row=1, column=3).value
    if isinstance(raw, (datetime, date)):
        d = raw.date() if isinstance(raw, datetime) else raw
        return d.isoformat()
    if isinstance(raw, str):
        # Try common formats.
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw.strip(), fmt).date().isoformat()
            except ValueError:
                continue
    return None


# Vehicle codes seen in the bottom-matrix HEADER row across the workbook.
# Used as the fall-back matrix-detection signal when the "Vehicle Earnings
# UK Deliveries" title cell has been blanked / typo'd (e.g. Wed_WK10_26 just
# has "d1" in B47).
KNOWN_VEHICLE_TOKENS: set[str] = {
    "B7MLC", "B12MLC", "B14MLC", "B15MLC", "B16MLC", "B18MLC", "B20MLC",
    "C2MLC", "C6MLC", "C12MLC", "C20MLC",
    "D1MLC", "D2MLC", "E1MLC",
    "WX17WSU", "X24CAL",
    "MLCD",                  # the typo'd column header; treat as a token here
                             # so it counts toward the matrix-row score
    "RENTAL", "SUBBY",
}


def find_matrix_header_row(ws) -> Optional[int]:
    """
    Return the row index of the matrix's HEADER row (the row containing the
    list of vehicle codes — C12MLC, B16MLC, etc.). Two strategies:

      1) The literal "Vehicle Earnings" label sits one row above the header.
      2) Otherwise, score every row in the matrix-search band by how many of
         its cells contain a known vehicle code; pick the highest scorer.

    Falls back to None if neither yields a confident match.
    """
    # Strategy 1: explicit label.
    for r in range(40, 80):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and "vehicle earnings" in v.lower():
            return r + 1   # header sits one row below the title

    # Strategy 2: vehicle-token scoring.
    best_row: Optional[int] = None
    best_count = 0
    for r in range(40, 65):
        count = 0
        for c in range(2, 26):
            val = ws.cell(row=r, column=c).value
            if isinstance(val, str):
                tok = val.strip().upper()
                if tok in KNOWN_VEHICLE_TOKENS:
                    count += 1
        if count > best_count:
            best_count = count
            best_row = r
    return best_row if best_count >= 5 else None


def matrix_total_for_index(ws, matrix_values_start: Optional[int], leg_index: int) -> float:
    """
    Sum the matrix row corresponding to a leg index (0-based, where leg_index=0
    means the first leg row in the legs section). `matrix_values_start` is the
    row of the FIRST £ row (one below the vehicle-codes header).
    """
    if matrix_values_start is None:
        return 0.0
    row_idx = matrix_values_start + leg_index
    total = 0.0
    # Sum a generous column range — the matrix is at most ~25 cols wide.
    for col in range(2, 30):
        total += safe_number(ws.cell(row=row_idx, column=col).value)
    return total


def is_leg_row(ws, r: int) -> bool:
    """A row counts as a leg if any of Collection/Vehicle/Customer is set."""
    a = safe_str(ws.cell(row=r, column=COL_COLLECTION).value)
    n = safe_str(ws.cell(row=r, column=COL_VEHICLE).value)
    v = safe_str(ws.cell(row=r, column=COL_CUSTOMER).value)
    return bool(a or n or v)


def extract_daily_sheet(ws, sheet_name: str) -> list[dict]:
    """Return a list of leg dicts for one daily sheet."""
    iso_date = header_date(ws)
    if not iso_date:
        # Some weeks may have a malformed header — skip silently.
        return []

    matrix_header_row = find_matrix_header_row(ws)
    matrix_values_start = (matrix_header_row + 1) if matrix_header_row else None

    legs: list[dict] = []
    leg_index = 0
    # Header row is row 19 (Collection/Delivery labels). Data starts row 20.
    # Stop scanning at the matrix HEADER row (the row of vehicle codes —
    # never a leg). When matrix detection fails, fall back to row 46 (the
    # earliest header row seen across the workbook), which is safer than
    # scanning into the matrix area and producing phantom legs like MLCD.
    max_leg_row = matrix_header_row if matrix_header_row else 46
    for r in range(20, max_leg_row):
        if not is_leg_row(ws, r):
            leg_index += 1
            continue

        # Walk the row's cells once for fill detection and field extraction.
        row_cells = list(ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=22))[0]
        run_type = detect_run_type(row_cells) or "regular"

        from_pc       = safe_str(ws.cell(row=r, column=COL_COLLECTION).value)
        to_pc         = safe_str(ws.cell(row=r, column=COL_DELIVERY).value)
        day_idx       = parse_int(ws.cell(row=r, column=COL_DAY_INDEX).value)
        day_count     = parse_int(ws.cell(row=r, column=COL_DAY_COUNT).value)
        factory       = safe_str(ws.cell(row=r, column=COL_FACTORY).value)
        booking_time  = safe_str(ws.cell(row=r, column=COL_BOOKING_TIME).value)
        # Normalise: case-fold (uppercase), trim, collapse whitespace.
        # Catches `B12MLc` → `B12MLC`, `Griffin` → `GRIFFIN`, etc.
        vehicle       = normalise_upper(safe_str(ws.cell(row=r, column=COL_VEHICLE).value))
        subby_driver  = safe_str(ws.cell(row=r, column=COL_SUBBY_DRIVER).value)
        subby_cost    = safe_number(ws.cell(row=r, column=COL_SUBBY_COST).value)
        trailer_num   = normalise_upper(safe_str(ws.cell(row=r, column=COL_TRAILER_NUM).value))
        trailer_drop  = safe_str(ws.cell(row=r, column=COL_TRAILER_DROP).value)
        reference     = safe_str(ws.cell(row=r, column=COL_REFERENCE).value)
        customer      = normalise_upper(safe_str(ws.cell(row=r, column=COL_CUSTOMER).value))

        # **Filter**: only legs with a registration assigned get imported.
        # Forward-planning placeholders (no vehicle yet) are skipped per the
        # operator's directive — those will be re-entered in the portal as
        # they get planned.
        if not vehicle:
            leg_index += 1
            continue

        revenue = matrix_total_for_index(ws, matrix_values_start, leg_index)

        legs.append({
            "sourceSheet": sheet_name,
            "sourceRow": r,
            "date": iso_date,
            "runType": run_type,
            "fromPostcode": from_pc,
            "toPostcode": to_pc,
            "dayIndex": day_idx,
            "dayCount": day_count if (day_idx and day_count) else None,
            "factory": factory or None,
            "bookingTime": booking_time or None,
            "vehicle": vehicle,
            "subbyDriver": subby_driver or None,
            "subbyCost": subby_cost if subby_cost > 0 else None,
            "trailerNumber": trailer_num or None,
            # The spreadsheet writes "MV" (Moreton Vallance) etc. as a free-text
            # marker rather than a strict bool. We treat any non-empty value
            # as "yes, dropped".
            "trailerDropped": bool(trailer_drop),
            "trailerDroppedNote": trailer_drop or None,
            "reference": reference or None,
            "customer": customer,
            "revenue": revenue,
            # billable + invoiceStatus are filled in by the Invoicing-sheet pass.
            "billable": False,
            "invoiceStatus": "open",
        })
        leg_index += 1
    return legs


# ── Invoicing sheet cross-reference ──────────────────────────────────────────

WEEKDAY_NAMES = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}


def extract_invoicing_sheet(ws) -> list[dict]:
    """
    Walks the Invoicing_WK<n>_<yy> sheet. Returns a list of dicts:
      { date, customer, fromPostcode, toPostcode, vehicle, reference,
        price, bill, status }
    Layout (per our earlier dump):
      A=Collection, B=Delivery, C=Customer, D=Factory, E=Price, F=Vehicle,
      G=Reference, ..., S=Description, T=Bill?, U=Status, V=XeroTag
    Date "header" rows have weekday name in A and the date in B.
    """
    rows: list[dict] = []
    cur_date: Optional[str] = None

    last_row = ws.max_row
    for r in range(1, last_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value

        # Date-header row: A is a weekday name (Monday..Sunday), B is a date.
        if isinstance(a, str) and a.strip().lower() in WEEKDAY_NAMES:
            if isinstance(b, (datetime, date)):
                d = b.date() if isinstance(b, datetime) else b
                cur_date = d.isoformat()
            continue

        if cur_date is None:
            continue

        from_pc  = safe_str(a)
        to_pc    = safe_str(b)
        customer = safe_str(ws.cell(row=r, column=3).value)
        # factory   = safe_str(ws.cell(row=r, column=4).value)
        price    = safe_number(ws.cell(row=r, column=5).value)
        vehicle  = safe_str(ws.cell(row=r, column=6).value)
        ref      = safe_str(ws.cell(row=r, column=7).value)
        bill     = ws.cell(row=r, column=20).value   # T
        status   = safe_str(ws.cell(row=r, column=21).value)  # U

        # Skip blank rows.
        if not from_pc and not to_pc and not customer and price == 0 and not vehicle:
            continue

        rows.append({
            "date": cur_date,
            "fromPostcode": from_pc,
            "toPostcode": to_pc,
            "customer": customer,
            "vehicle": vehicle,
            "reference": ref,
            "price": price,
            "bill": is_truthy(bill),
            "status": status.lower(),
        })
    return rows


def is_truthy(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    return s in ("true", "yes", "y", "1", "x", "✓")


def normalise(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().upper())


def merge_invoicing_into_legs(legs: list[dict], invoicing: list[dict]) -> dict:
    """
    Match Invoicing rows back to legs by (date, customer, from, to, vehicle).
    Mutates `legs` in place. Returns a small stats dict for reporting.
    """
    by_key: dict[tuple, list[dict]] = {}
    for inv in invoicing:
        key = (
            inv["date"],
            normalise(inv["customer"]),
            normalise(inv["fromPostcode"]),
            normalise(inv["toPostcode"]),
            normalise(inv["vehicle"]),
        )
        by_key.setdefault(key, []).append(inv)

    matched = 0
    sent = 0
    for leg in legs:
        key = (
            leg["date"],
            normalise(leg["customer"]),
            normalise(leg["fromPostcode"]),
            normalise(leg["toPostcode"]),
            normalise(leg["vehicle"]),
        )
        candidates = by_key.get(key)
        if not candidates:
            continue
        # If multiple invoicing rows share the same key, prefer one that
        # matches on revenue.
        chosen = candidates[0]
        if len(candidates) > 1:
            for c in candidates:
                if abs(c["price"] - leg["revenue"]) < 0.01:
                    chosen = c
                    break

        leg["billable"] = bool(chosen["bill"])
        if chosen["status"] == "sent":
            leg["invoiceStatus"] = "sent"
            leg["billable"] = True
            sent += 1
        elif chosen["bill"]:
            leg["invoiceStatus"] = "billable"
        # Use Invoicing-sheet price as a tiebreaker if the matrix gave 0.
        if leg["revenue"] == 0 and chosen["price"] > 0:
            leg["revenue"] = chosen["price"]
        # Carry the customer-side reference if the leg didn't have one.
        if not leg["reference"] and chosen["reference"]:
            leg["reference"] = chosen["reference"]
        matched += 1

    return {
        "invoicingRows": len(invoicing),
        "legsMatched": matched,
        "legsSent": sent,
    }


# ── Workbook walker ──────────────────────────────────────────────────────────

WEEK_PATTERN = re.compile(r"^(?P<day>Mon|Tue|Wed|Thu|Fri|Sat|Sun)_WK(?P<wk>\d+)_(?P<yy>\d+)$")


def list_week_packs(wb) -> list[tuple[int, int]]:
    """Return sorted list of (yy, wk) pairs present in the workbook."""
    packs: set[tuple[int, int]] = set()
    for name in wb.sheetnames:
        m = WEEK_PATTERN.match(name)
        if m:
            packs.add((int(m["yy"]), int(m["wk"])))
    return sorted(packs)


def extract_week(wb, yy: int, wk: int) -> tuple[list[dict], dict]:
    """Extract one week's legs and merge invoicing state. Returns (legs, stats)."""
    suffix = f"_WK{wk}_{yy:02d}"
    legs: list[dict] = []
    for day in ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"):
        sheet_name = f"{day}{suffix}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        legs.extend(extract_daily_sheet(ws, sheet_name))

    invoicing = []
    inv_sheet_name = f"Invoicing{suffix}"
    if inv_sheet_name in wb.sheetnames:
        invoicing = extract_invoicing_sheet(wb[inv_sheet_name])

    merge_stats = merge_invoicing_into_legs(legs, invoicing)
    stats = {
        "year": 2000 + yy,
        "week": wk,
        "legCount": len(legs),
        "outbound": sum(1 for l in legs if l["runType"] == "regular"),
        "backload": sum(1 for l in legs if l["runType"] == "backload"),
        "billable": sum(1 for l in legs if l["billable"]),
        "sent": sum(1 for l in legs if l["invoiceStatus"] == "sent"),
        "totalRevenue": round(sum(l["revenue"] for l in legs), 2),
        **merge_stats,
    }
    return legs, stats


# ── CLI ──────────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(description="Extract MLC Master Planner into JSON")
    p.add_argument("--path", default=DEFAULT_PATH, help="Path to the .xlsm")
    p.add_argument("--week", type=int, help="ISO week (1..53)")
    p.add_argument("--year", type=int, help="2-digit year (e.g. 25 for 2025)")
    p.add_argument("--all", action="store_true", help="Process every weekly pack")
    p.add_argument("--out", required=True, help="Output JSON path")
    args = p.parse_args()

    if not args.all and not (args.week and args.year is not None):
        p.error("Provide --all OR both --week and --year")

    print(f"Loading {args.path} …")
    wb = openpyxl.load_workbook(args.path, data_only=True, keep_vba=False, read_only=False)

    if args.all:
        packs = list_week_packs(wb)
    else:
        packs = [(args.year, args.week)]
    print(f"Processing {len(packs)} weekly pack(s)\n")

    all_legs: list[dict] = []
    all_stats: list[dict] = []
    for (yy, wk) in packs:
        legs, stats = extract_week(wb, yy, wk)
        all_legs.extend(legs)
        all_stats.append(stats)
        print(
            f"  WK{wk:02d}_{yy:02d}  "
            f"legs={stats['legCount']:3d}  "
            f"out={stats['outbound']:3d}  back={stats['backload']:3d}  "
            f"bill={stats['billable']:3d}  sent={stats['sent']:3d}  "
            f"rev=£{stats['totalRevenue']:>9,.2f}"
        )

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps({
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "sourceFile": args.path,
        "stats": all_stats,
        "legs": all_legs,
    }, indent=2, default=str), encoding="utf-8")

    total_rev = sum(s["totalRevenue"] for s in all_stats)
    total_sent = sum(s["sent"] for s in all_stats)
    print()
    print(f"Wrote {len(all_legs)} legs across {len(packs)} weeks → {out_path}")
    print(f"Totals: £{total_rev:,.2f} revenue · {total_sent} sent invoices")
    return 0


if __name__ == "__main__":
    sys.exit(main())
